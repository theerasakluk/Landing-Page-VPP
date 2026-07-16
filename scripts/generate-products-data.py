from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = Path(
    r"C:\Users\Theerasak\VPP Products\outputs\019f64c7\รายการสินค้าVVP_พร้อมรายละเอียดและรูปสินค้า_ฉบับสมบูรณ์.xlsx"
)
OUTPUT = ROOT / "data" / "products.ts"


CATEGORIES = [
    {
        "slug": "office-supplies",
        "name": "อุปกรณ์สำนักงานเบ็ดเตล็ด",
        "description": "เครื่องใช้ประจำสำนักงาน อุปกรณ์โต๊ะทำงาน และของใช้ทั่วไป",
    },
    {
        "slug": "office-electronics",
        "name": "อุปกรณ์สำนักงานอิเล็กทรอนิกส์",
        "description": "เครื่องคิดเลข โทรศัพท์ ปลั๊กไฟ และอุปกรณ์ไฟฟ้าในสำนักงาน",
    },
    {
        "slug": "office-furniture",
        "name": "เฟอร์นิเจอร์สำนักงาน",
        "description": "โต๊ะ เก้าอี้ ตู้เอกสาร ชั้นวาง และเฟอร์นิเจอร์สำหรับออฟฟิศ",
    },
    {
        "slug": "tools-equipment",
        "name": "อุปกรณ์และเครื่องมือ",
        "description": "เครื่องมือช่าง อุปกรณ์ซ่อมบำรุง อุปกรณ์เซฟตี้ และของใช้หน้างาน",
    },
    {
        "slug": "pantry-cleaning",
        "name": "เครื่องดื่มเครื่องใช้และผลิตภัณฑ์อื่นๆ",
        "description": "สินค้าแคนทีน ผลิตภัณฑ์ทำความสะอาด และของใช้ประจำอาคาร",
    },
    {
        "slug": "stationery-paper",
        "name": "อุปกรณ์เครื่องเขียน และผลิตภัณฑ์กระดาษ",
        "description": "ปากกา แฟ้ม สมุด กระดาษ แบบฟอร์ม และอุปกรณ์งานเอกสาร",
    },
    {
        "slug": "computer-it",
        "name": "ผลิตภัณฑ์สำหรับคอมพิวเตอร์และไอทีต่างๆ",
        "description": "อุปกรณ์ต่อพ่วง หมึกพิมพ์ สายเคเบิล และสินค้าไอทีสำนักงาน",
    },
    {
        "slug": "solar-rooftop",
        "name": "ผลิตภัณฑ์สำหรับโซล่าร์รูฟท็อป",
        "description": "อุปกรณ์พลังงาน ระบบติดตั้ง และสินค้าเกี่ยวกับ solar rooftop",
    },
]


SUBCATEGORY_RULES = {
    "stationery-paper": [
        ("books-learning", "หนังสือและสื่อการเรียน", ["หนังสือ", "แบบเรียน", "ข้อสอบ", "นิทาน", "วรรณ", "การศึกษา"]),
        ("paper-files", "กระดาษ สมุด และแฟ้ม", ["กระดาษ", "สมุด", "แฟ้ม", "เอกสาร", "ซอง", "ปก"]),
        ("writing-tools", "ปากกา ดินสอ และเครื่องเขียน", ["ปากกา", "ดินสอ", "ยางลบ", "ไม้บรรทัด", "สี", "marker", "เมจิก"]),
        ("creative-supplies", "งานศิลปะและสื่อสร้างสรรค์", ["ระบาย", "พู่กัน", "โปสเตอร์", "กาว", "สติ๊กเกอร์", "สติกเกอร์", "งานประดิษฐ์"]),
    ],
    "office-supplies": [
        ("uniform-event", "เครื่องแต่งกายและกิจกรรม", ["ชุด", "รปภ", "ปลอกแขน", "นกหวีด", "ผ้า", "เกียรติบัตร"]),
        ("personal-care", "ของใช้ส่วนตัวและความงาม", ["สเปรย์", "กิ๊บ", "รองพื้น", "แชมพู", "สบู่", "ครีม"]),
        ("document-office", "งานเอกสารและอุปกรณ์สำนักงาน", ["กรอบ", "คลิป", "ตรายาง", "เทป", "ซอง", "กล่อง", "เอกสาร"]),
    ],
    "tools-equipment": [
        ("cleaning-maintenance", "ทำความสะอาดและดูแลพื้นที่", ["แปรง", "ถุงดำ", "ผงซัก", "สก๊อต", "น้ำยา", "ไม้ถู", "ซับน้ำ", "ขัด"]),
        ("repair-tools", "เครื่องมือช่างและซ่อมบำรุง", ["ค้อน", "ไขควง", "ประแจ", "สว่าน", "คีม", "ใบเลื่อย", "เทปพัน"]),
        ("site-materials", "วัสดุก่อสร้างและหน้างาน", ["หิน", "ทราย", "ปูน", "ยางมะตอย", "เหล็ก", "ท่อ", "สีทา"]),
    ],
}


def normalize(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def slug_for_category(category_name: str) -> str | None:
    for category in CATEGORIES:
        if category["name"] == category_name:
            return category["slug"]
    return None


def infer_subcategory(category_slug: str, name: str, detail: str) -> str:
    haystack = f"{name} {detail}".lower()
    for subcategory_slug, _label, keywords in SUBCATEGORY_RULES.get(category_slug, []):
        if any(keyword.lower() in haystack for keyword in keywords):
            return subcategory_slug
    return "other" if category_slug in SUBCATEGORY_RULES else ""


def main() -> None:
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    rows = worksheet.iter_rows(values_only=True)
    headers = [normalize(value) for value in next(rows)]
    header_index = {header: index for index, header in enumerate(headers)}

    products = []
    for row in rows:
        if not row:
            continue

        category_name = normalize(row[header_index["หมวดหมู่สินค้า"]])
        category_slug = slug_for_category(category_name)
        product_name = normalize(row[header_index["รายการสินค้า"]])
        if not category_slug or not product_name:
            continue

        detail = normalize(row[header_index["รายละเอียดสินค้า"]])
        products.append(
            {
                "id": len(products) + 1,
                "categorySlug": category_slug,
                "subcategorySlug": infer_subcategory(category_slug, product_name, detail),
                "code": normalize(row[header_index["รหัสสินค้า"]]),
                "name": product_name,
                "detail": detail,
            }
        )

    subcategory_meta = {
        category_slug: [{"slug": slug, "label": label} for slug, label, _keywords in rules] + [{"slug": "other", "label": "อื่นๆ"}]
        for category_slug, rules in SUBCATEGORY_RULES.items()
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    content = "\n".join(
        [
            "export type ProductCategory = {",
            "  slug: string;",
            "  name: string;",
            "  description: string;",
            "};",
            "",
            "export type ProductSubcategory = {",
            "  slug: string;",
            "  label: string;",
            "};",
            "",
            "export type Product = {",
            "  id: number;",
            "  categorySlug: string;",
            "  subcategorySlug: string;",
            "  code: string;",
            "  name: string;",
            "  detail: string;",
            "};",
            "",
            f"export const categoryMeta = {json.dumps(CATEGORIES, ensure_ascii=False, indent=2)} as const satisfies readonly ProductCategory[];",
            "",
            f"export const subcategoryMeta = {json.dumps(subcategory_meta, ensure_ascii=False, indent=2)} as const satisfies Record<string, readonly ProductSubcategory[]>;",
            "",
            f"export const products = {json.dumps(products, ensure_ascii=False, indent=2)} as const satisfies readonly Product[];",
            "",
            "export function getCategoryBySlug(slug: string) {",
            "  return categoryMeta.find((category) => category.slug === slug);",
            "}",
            "",
            "export function getProductsByCategorySlug(slug: string) {",
            "  return products.filter((product) => product.categorySlug === slug);",
            "}",
            "",
            "export function getSubcategoryGroups(slug: string) {",
            "  return subcategoryMeta[slug] ?? [];",
            "}",
            "",
        ]
    )
    OUTPUT.write_text(content, encoding="utf-8")

    counts = {}
    for product in products:
        counts[product["categorySlug"]] = counts.get(product["categorySlug"], 0) + 1
    print(json.dumps({"products": len(products), "counts": counts}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
